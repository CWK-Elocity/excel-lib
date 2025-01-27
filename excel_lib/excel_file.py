import openpyxl.utils
import openpyxl.utils.exceptions
import pandas as pd
import openpyxl
import io
import zipfile
import env

def file_to_io_stream(path):
    with open(path, "rb") as file:
        file_stream = io.BytesIO(file.read())
    return file_stream

def is_match(value1, value2):
    """Compares two values, stripping strings if applicable."""
    if isinstance(value1, str) and isinstance(value2, str):
        return value1.strip() == value2.strip()
    return value1 == value2

def find_row_for_key(key):
    """Finds the row index for a given key in the worksheet."""
    for row_index, value in self.worksheet.iloc[:, 1].items():
        if pd.notna(value) and is_match(value, key):
            return row_index
    return -1

class ExcelFile:
    def __init__(self, file_stream):
        """Creates an ExcelFile object that will deal with data processing
            Validates file and checks for images

        Args:
            file_stream (io_stream): an excel file in form of io_stream
        """
        self.worksheet_count = 0
        try:
            workbook = self._validate_excel_file(file_stream)
            self.non_cell_objects = self._check_for_non_cell_objects(workbook)
        except Exception as e:
            print(f"Unecpected error occured during loading file into openpyxl: {e}")
        self.worksheet = pd.read_excel(file_stream)
    
    def _validate_excel_file(self, file_stream):
        """Validates excel file and checks how many worksheets there are in file

        Args:
            file_stream (io_stream): an excel file in form of io_stream

        Raises:
            ValueError: If it did not pass validation

        Returns:
            object: first workbook from excel file
        """
        try:
            file_stream.seek(0)
            workbook = openpyxl.load_workbook(file_stream)
            self.worksheet_count = len(workbook.sheetnames)
            return workbook
        except openpyxl.utils.exceptions.InvalidFileException:
            raise ValueError("It is not a valid excel file")
        
    def _check_for_non_cell_objects(self, openpyxl_workbook_instance):
        """Checks if there are any objects like images or charts etc. 
        or other things that are not in cells

        Args:
            openpyxl_workbook_instance (object): workbook from excel file

        Returns:
            lsit: list of non cell objetct
        """
        workbook = openpyxl_workbook_instance
        non_cell_objects = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            if worksheet._images:
                for image in worksheet._images:
                    if image.anchor._from:
                        anchor = image.anchor._from
                        anchor_info = f"Image anchored at cell {anchor.col + 1}{anchor.row + 1} in sheet '{sheet_name}'."
                    else:
                        anchor_info = f"Image not anchored to any cell in sheet '{sheet_name}'."
                    non_cell_objects.append(anchor_info)

            if worksheet._charts:
                for chart in worksheet._charts:
                    non_cell_objects.append(f"Chart found in sheet '{sheet_name}'.")
        
        return non_cell_objects
    
    """
    obsolete
    def _check_for_images_in_archive(self, file_stream):
        
        images_found = []

        file_stream_copy = io.BytesIO(file_stream.getvalue())

        with zipfile.ZipFile(file_stream_copy, 'r') as zip_ref:
            image_files = [file for file in zip_ref.namelist() if file.startswith("xl/media/")]
            if image_files:
                for image_file in image_files:
                    images_found.append(f"Image found: {image_file}")
            else:
                images_found.append("No images found in xl/media/ folder")

        return images_found

    def get_non_cell_objects_info(self):
        if self.non_cell_objects:
            return "\n".join(self.non_cell_objects)
        return "No non-cell objects detected."
    
    def get_sheet_names(self):
        return self.workbook.sheetnames
    """

    def _identify_sections(self):
        """Identify sections based on first column in workbook

        Returns:
            list: list of section names
        """
        sections = {}
        current_section = None
        for row_index, value in enumerate(self.worksheet.iloc[:, 0]):
            if isinstance(value, str) and value.isupper():
                if current_section:
                    sections[current_section][1] = row_index
                if isinstance(value, str) and isinstance(value, str):
                    current_section = value.strip()
                else:
                    current_section = value
                sections[current_section] = [row_index, None]

        if current_section:
            sections[current_section][1] = self.worksheet.iloc[:, 0].last_valid_index() + 1
        return sections
    
    def create_template_structure(self):
        """Based on example file creates template with row numbers where the information should be

        Returns:
            structure/json object: dictionary with name of information: rowa frouped by sections
        """
        template_structure = {
            "takeover": {
                "global_data": {},
                "contact_person": None,
                "responsible_person": None
            },
            "stations": {}
        }

        # Identify sections
        sections = self._identify_sections()

        # Find the key corresponding to global_data
        takeover_divider_key = next((key for key in env.SECTION_STATION_TAKEOVER_DIVIDER if key in sections), None)

        if takeover_divider_key:
            global_data = {}
            for row_index, row in self.worksheet.iloc[:sections[takeover_divider_key][0], :2].iterrows():
                value, key = row
                if pd.notna(key) and pd.notna(value):
                    global_data[key] = row_index
            template_structure["takeover"]["global_data"] = global_data

        # Find the key corresponding to contact_person
        contact_person_key = next((key for key in env.SECTION_CONTACT_PERSON if key in sections), None)
        if contact_person_key:
            contact_person = {}
            for row_index, row in self.worksheet.iloc[sections[contact_person_key][0]:sections[contact_person_key][1], :2].iterrows():
                value, key = row
                if pd.notna(key) and pd.notna(value):
                    contact_person[key] = row_index
            template_structure["takeover"]["contact_person"] = contact_person

        # Find the key corresponding to responsible_person
        responsible_person_key = next((key for key in env.SECTION_RESPONSIBLE_PERSON if key in sections), None)
        if responsible_person_key:
            responsible_person = {}
            for row_index, row in self.worksheet.iloc[sections[responsible_person_key][0]:sections[responsible_person_key][1], :2].iterrows():
                value, key = row
                if pd.notna(key) and pd.notna(value):
                    responsible_person[key] = row_index
            template_structure["takeover"]["responsible_person"] = responsible_person

        # Station data (from "STACJA ŁADOWANIA - DANE" section and onwards)
        station_structure = {}
        for section, section_range in sections.items():
            station_data = {}
            for row_index in range(section_range[0], section_range[1]):
                key = self.worksheet.iat[row_index, 1]
                if pd.notna(key):
                    station_data[key] = row_index
            station_structure[section] = station_data
        template_structure["stations"]=station_structure

        return template_structure

    def retrive_stations(self):
        """_summary_

        Returns:
            _type_: _description_
        """
        if self.discarded_data_info:
            pass
        else:
            self.discarded_data_info = []
        stations = []
        length = self.comparison_template.shape[0]
        for loop_index, (column_name, column_data) in enumerate(self.worksheet.iloc[:, 2:].items()):
            station_data = self.comparison_template.iloc[:, 1:].copy().values.tolist()
            for form_index, value, index in self.comparison_template.itertuples(index=False):
                if index is not -1:
                    station_data[form_index][1] = column_data.iat[index]

            nones = pd.isna([row[1] for row in station_data]).sum()

            if length - nones <= 3:
                self.discarded_data_info.append(f"Column {loop_index + 2} not taken int account. Too little data.")
            else:
                stations.append(station_data)

        return stations
    
    def _update_rows_in_structure(self, data_section):
        """Checks if the value is in the same row in file and in template.
        Otherwise looks for that specific value in all rows, and if found then updates row number.

        Args:
            data_section (dictionary): section of whole data

        Returns:
            dictionary: updated section
        """

        updated_section = {}
        for key, expected_row in data_section.items():
            actual_label = self.worksheet.iloc[expected_row, 1] if expected_row < len(self.worksheet) else None

            if pd.notna(actual_label) and is_match(actual_label, key):
                updated_section[key] = expected_row
            else:
                updated_section[key] = find_row_for_key(key)

        return updated_section

    def compare_structure_with_file(self, template):
        """Compares actual working file with give template to obtain rows that will be used to determine value

        Args:
            template (dictionary): a dict containg template wich user wants to use

        Returns:
            dcit: similar to template but updated for that file
        """
        updated_structure = {
            "takeover": {
                "global_data": {},
                "contact_person": None,
                "responsible_person": None
            },
            "stations": {}
        }

        for section_name, takeover_section in template["takeover"].items():
            updated_structure["takeover"][section_name] = self._update_rows_in_structure(takeover_section)

        for section_name, station_section in template["stations"].items():
            updated_structure["stations"][section_name] = self._update_rows_in_structure(station_section)

        return updated_structure


    def create_data_structure_from_template(self, template):
        """Gather data from file based on template and structurize them in one dict object

        Args:
            template (dictionary): a dict containg template wich user wants to use

        Returns:
            dict: containg ale data categorised into sections
        """

        data_structure = self.compare_structure_with_file(template)

        collected_takeover_structures = []

        for column in range(2, self.worksheet.shape[1]):
            # Take global data from column
            current_global_data = {
                key: self.worksheet.iloc[row, column] if row < len(self.worksheet) else None for key, row in data_structure["takeover"]["global_data"].items()
            }

            # Check whether there is a struct with this global data
            matching_group = None
            for group in collected_takeover_structures:
                if group["global_data"] == current_global_data:
                    matching_group = group
                    break

            # If there is none, create new one
            if not matching_group:
                matching_group = {
                    "global_data": current_global_data,
                    "contact_person": None,
                    "responsible_person": None,
                    "stations": []
                }
                collected_takeover_structures.append(matching_group)

            # Compare contact person
            current_contact_person = {
                key: self.worksheet.iloc[row, column] if row < len(self.worksheet) else None for key, row in data_structure["takeover"]["contact_person"].items()
            }
            if matching_group["contact_person"] is None:
                matching_group["contact_person"] = current_contact_person
            elif matching_group["contact_person"] != current_contact_person:
                matching_group["contact_person"] = "Dla każdej stacji inna"

            # Compare responsible person
            current_responsible_person = {
                key: self.worksheet.iloc[row, column] if row < len(self.worksheet) else None for key, row in data_structure["takeover"]["responsible_person"].items()
            }
            if matching_group["responsible_person"] is None:
                matching_group["responsible_person"] = current_responsible_person
            elif matching_group["responsible_person"] != current_responsible_person:
                matching_group["responsible_person"] = "Dla każdej stacji inna"

            # Add station data
            station_data = {
                section: {
                    field: self.worksheet.iloc[row, column] if row < len(self.worksheet) else None
                    for field, row in fields.items()
                }
                for section, fields in data_structure["stations"].items()
            }
            matching_group["stations"].append(station_data)

        return collected_takeover_structures

